"""
Export leads to CSV and / or styled Excel.
"""

import csv
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console

from models import Lead
import config

console = Console()


def _ensure_dir():
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)


def _filename(query: str, ext: str) -> str:
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in query)
    safe = safe.strip().replace(" ", "_")[:50]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(config.OUTPUT_DIR, f"{safe}_{ts}.{ext}")


# ═══════════════════════════════════════════════════════════════════════
#  CSV
# ═══════════════════════════════════════════════════════════════════════

def export_to_csv(leads: list[Lead], query: str) -> str:
    _ensure_dir()
    path = _filename(query, "csv")
    if not leads:
        console.print("[yellow]  ⚠ No leads to export[/]")
        return ""

    rows = [l.to_dict() for l in leads]
    headers = list(rows[0].keys())

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)

    console.print(f"[bold green]  📄 CSV saved →[/] {path}")
    return path


# ═══════════════════════════════════════════════════════════════════════
#  Excel (styled)
# ═══════════════════════════════════════════════════════════════════════

def export_to_excel(leads: list[Lead], query: str) -> str:
    _ensure_dir()
    path = _filename(query, "xlsx")
    if not leads:
        console.print("[yellow]  ⚠ No leads to export[/]")
        return ""

    rows = [l.to_dict() for l in leads]
    headers = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Styles
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side("thin", "D1D5DB"), right=Side("thin", "D1D5DB"),
        top=Side("thin", "D1D5DB"), bottom=Side("thin", "D1D5DB"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F4F6")

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    # Data rows
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h, ""))
            c.alignment, c.border = cell_align, border
            if ri % 2 == 0:
                c.fill = alt_fill

    # Auto-width
    for ci, h in enumerate(headers, 1):
        max_len = len(h)
        for ri in range(2, len(rows) + 2):
            val = str(ws.cell(row=ri, column=ci).value or "")
            max_len = max(max_len, min(len(val), 55))
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 4

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)

    console.print(f"[bold green]  📊 Excel saved →[/] {path}")
    return path


# ═══════════════════════════════════════════════════════════════════════
#  Dispatcher
# ═══════════════════════════════════════════════════════════════════════

def export_leads(leads: list[Lead], query: str) -> list[str]:
    console.print(f"\n[bold cyan]💾 Exporting {len(leads)} leads…[/]\n")
    files: list[str] = []
    if config.OUTPUT_FORMAT in ("csv", "both"):
        p = export_to_csv(leads, query)
        if p:
            files.append(p)
    if config.OUTPUT_FORMAT in ("excel", "both"):
        p = export_to_excel(leads, query)
        if p:
            files.append(p)
    return files
